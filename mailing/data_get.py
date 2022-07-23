import openpyxl as xls
import os
import sys
from openpyxl.workbook import workbook
from make_logs.logger import logging, wipe
from .path_consts import BASIC_XL_PATH ,BASIC_Q_PATH ,BASIC_S_PATH
#this class inits the d8icts i need in order to start sending 
#our excel stracture should be like [anagnoristiko, email, people_name, afm]
class getter:
    def __init__(self, path_xl=BASIC_XL_PATH, path_s=BASIC_S_PATH, path_q=BASIC_Q_PATH):
        self.path_xl= path_xl
        self.path_s=path_s
        self.path_q=path_q
        self.sending=dict() #anagnoristiko : email
        self.exist=dict() #anagnoristiko : 0 #should be anagnoristiko:2 in order to be okay and exist in both files
        self.willSend=dict() #anagnoristiko:false #should be true in order for a222n email to be sent there
        self.total_issues=0 #used to print that everything is okay if no issues are found
        self._run()

    #used to print if everything is fine and then logs it
    def All_ok(self):
        if self.total_issues ==0:
            msg= "there are 0 issues while chekcing files"
            print(msg)
            logging(msg,True)

    #gets all data needed from the excel and stores them in the dicts
    def _get_data(self):
        workbook= xls.load_workbook(self.path_xl)
        sheet=workbook.worksheets[0]
        for i in range(2,sheet.max_row+1):
            anagnoristiko= sheet.cell(column=1, row=i).value
            email=sheet.cell(column= 2, row= i).value
            self.sending[anagnoristiko]=email
            self.exist[anagnoristiko]= 0
            self.willSend[anagnoristiko]= False
    
    #checks if all anagnoristika exist 
    def _check_files(self,path: str, start: str, end: str):
        dir_list= os.listdir(path)
        issues=0
        good=True
        for key in self.sending:
            file= start+key+end
            if file in dir_list:
                self.exist[key]+=1  #checks if file exists in the dir or else it prints it doesn't prints number of issues at the end and logs everything
            else:
                msg= f'{key} does not exist in {path}'
                print(msg)
                logging(msg,False)
                issues+=1 
        msg=f'number of issues :{issues} in path {path}'
        print(msg)
        self.total_issues+=issues
        if issues>0:
            good=False
        logging(msg,good)
        

         
    #this deals with willSend dict checks if there wont be sent some emails  
    def _all_is_ok(self):
        for key in self.exist:
            if self.exist[key]!=2:
                msg= f"file with anagnoristiko: {key} exists {self.exist[key]} times, no email will be sent to {self.sending[key]}"
                print(msg)
                logging(msg, False)
                self.total_issues+=1
            else:
                self.willSend[key]=True

    def _run(self):
        wipe() #wipes data from the logs to start anew
        self._get_data()
        self._check_files(self.path_q,'Q','.pdf')
        self._check_files(self.path_s,'S','.pdf')
        self._all_is_ok()
        self.All_ok()


if __name__ == "__main__":
    feeder()
