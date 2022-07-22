from pickle import NONE
import random as ra
import os

from openpyxl.workbook import workbook
from name_consts import ALPHABET, MAILS
import openpyxl as xls


# creates all basic test files(txt and pdf)
class test_file_creator:

    # a function that creates and returns a -list with all the test file names
    def create_names(self, start: int, end: int):
        file_name = list()
        for i in range(start, end + 1, 1):
            file_name += ["Test_file_" + str(i)]
        return file_name

    # creates a random afm everytime its called
    def create_afm(self):
        afm = ""
        afm += str(ra.randint(0, 1))
        for i in range(7):
            afm += str(ra.randint(0, 9))
        return afm
    
    #makes a random anagnoristiko for the xlsx file
    def create_an(self):
        an=''
        length=9
        for i in range(length):
            if ra.randint(1,5) ==3: #adds a one in 5 chance that there will be a number in an instead of letter
                an+=str(ra.randint(0,9))
            else:
                letter=ra.choice(ALPHABET)
                if ra.randint(1,2)==1: #adds a 1 in 2 chance of letter to be capitalized
                    letter=letter.upper()
                an+=letter
        return an

    def make_name(self):
        name=''
        name+=ra.choice(ALPHABET).upper() #adds a capitalized first letter to the name it returns
        for i in range(ra.randint(4, 10)): #makes len of name randomized
            name+=ra.choice(ALPHABET)
        name+=' '
        for i in range(ra.randint(4, 10)): #makes len of last_name randomized
            name+=ra.choice(ALPHABET)    
        return name

    #this function makes xlsx files with data like (anagnoristiko,email,name_of_person,afm)
    #and then adds the data i want to it
    def make_xlsx_file(self,workbook , current_an: str, persons_name: str, current_afm: str, current_email: str, name_file: str,faulty: bool):
        names=['anagnoristiko','email','people_name','afm']
        data=[current_an, current_email, persons_name, current_afm]
        path= 'File_test/'+name_file+'.xlsx'
        if workbook==None:
            workbook= xls.Workbook()
            worksheet = workbook.worksheets[0]
            for index,current in enumerate(names):          #makes a new one with the top cells having the names i want
                cell=worksheet.cell(row=1, column=index+1 )
                cell.value= current
        else: worksheet= workbook.worksheets[0]             #this is extremely bad code     
        if (faulty and ra.randint(0, 999) == 500):  # adds 1/1000 chance to have a faulty afm
            letter = ra.choice(current_an)
            data[0] = current_an.replace(letter, "")
        t_row=worksheet.max_row
        if worksheet.cell(row=t_row, column=1).value!=None: t_row+=1                                                                                                                                   
        for index,current in enumerate(names):
            worksheet.cell(row= t_row, column=index+1).value=data[index]
        return workbook,path           

    # creates a random email everytime its called
    def _create_email(self):
        email = ""
        name_len = ra.randint(5, 15)
        for i in range(name_len):
            email += ra.choice(ALPHABET)
        email += "@"
        email += ra.choice(MAILS)
        return email

       # makes a direcotry that containts all the pdfs corresponding to a file
    def make_pdf_dir(self, current_index):
        name = "PDF_files/" + "pdf_files_" + str(current_index) + "/"
        namea= "PDF_files/" + "pdf_files_" + str(current_index) + "/items/" 
        nameq= "PDF_files/" + "pdf_files_" + str(current_index) + "/questions/"
        os.mkdir(name)
        os.mkdir(namea)
        os.mkdir(nameq)
        return name

    # makes an empty file with a .pdf extension so we can have something to send
    def make_pdf_file(self, current_afm, current_pdf_dir):
        try:
            with open(current_pdf_dir+'questions/' + "Q" + current_afm + ".txt", "a") as f:
                f.write("he")
            os.rename(current_pdf_dir+'questions/' + "Q" + current_afm +".txt", current_pdf_dir + "questions/Q" + current_afm +".pdf")
            with open (current_pdf_dir+'items/' + "S" + current_afm + ".txt", "a") as f: 
                 f.write("he")
            os.rename(current_pdf_dir+'items/' + "S" + current_afm +".txt", current_pdf_dir + "items/S" + current_afm +".pdf")

        except Exception as e:
            print(e,"im here")
        return current_afm

    # uses the functions to make the files i want to test if create_pdfs or create_txt is False no txt or pdf will be made
    # takes range of files(statr, end) and if i want the files to have faults or no(afm missing a letter)
    def run(
        self,
        start: int,
        end: int,
        numberOfPeople: int,
        create_pdfs: bool,
        create_txt: bool,
        faulty: bool,
        ):
        names = self.create_names(start, end)
        for i in range(end+1 - start):  # make (end - start) amount of files
            print(i, "now")
            workbook=None
            if create_pdfs:
                current_pdf_dir = self.make_pdf_dir(i+1)
            for j in range(numberOfPeople):  # with (numberOfPeople) amount of lines
                current_afm = self.create_afm()
                current_email = self._create_email()
                current_name= self.make_name()
                current_an=self.create_an()
                if create_txt:
                    workbook,path= self.make_xlsx_file(workbook,current_an,current_name,current_afm,current_email,names[i],faulty)
                if create_pdfs:
                    check_pdf= self.make_pdf_file(current_an, current_pdf_dir)
            workbook.save(path)

# function that creates the files needed
def Create_files(
    start=1, end=10, people=4000, faulty=False, create_pdfs=True, create_txt=True
    ):
    if (create_txt or create_pdfs): #only runs if a file is needed to be created
        test_file_creator().run(start, end, people, create_pdfs, create_txt, faulty)

#made so if run by its self it just makes all the files
if __name__ == "__main__":
    Create_files()







